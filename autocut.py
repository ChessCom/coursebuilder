#!/usr/bin/env python3
"""
autocut.py — Reads cut points from a Recording Schedule Excel and applies them
             to the matching sequence in Adobe Premiere Pro 2026 via ExtendScript.

Structure assumed:
  • ONE sequence per chapter (e.g. "4a) Italian 4.d3: Taunt of Destiny Declined")
  • The sequence contains individual video clips: Chapter4aPart1.mp4, Chapter4aPart2.mp4 …
  • Cut times in the Excel are relative to the START of each Part clip

Usage:
    python3 autocut.py "Chapter 4a Part2" "/path/to/03_Course"
    python3 autocut.py "Chapter 4a Part2"          # uses default course folder
"""

import sys
import os
import re
import json
import time
import urllib.request
import urllib.error
import datetime
import openpyxl

# ─── Default paths ────────────────────────────────────────────────────────────
DEFAULT_COURSE_FOLDER = (
    "/Users/raulmartinez/Desktop/chess.com/"
    "1.e4 e5 Gambit Repertoire (TBD!)/"
    "03_Course"
)
AUTOCUT_SERVER = "http://127.0.0.1:7788"
VERSION        = "0.29"


# ─── Name parsing ─────────────────────────────────────────────────────────────

def normalize(text: str) -> str:
    """Lowercase, strip all non-alphanumeric chars."""
    return re.sub(r"[^a-z0-9]", "", str(text).lower())


def parse_chapter_name(name: str) -> tuple[str | None, str | None]:
    """
    "Chapter 4a Part2"   → ("4a", "2")
    "Chapter4aPart 10"   → ("4a", "10")
    "Chapter 3a Part 1"  → ("3a", "1")
    Returns (chapter_id, part_number) or (None, None) if no match.
    """
    m = re.search(r"chapter\s*([0-9]+[a-z]*)\s*part\s*([0-9]+)", name, re.IGNORECASE)
    if m:
        return m.group(1).lower(), m.group(2)
    return None, None


# ─── Excel helpers ────────────────────────────────────────────────────────────

def find_excel(course_folder: str) -> str | None:
    """
    Search for 'Recording Schedule*.xlsx' starting at course_folder,
    then walking up parent directories (up to 4 levels) if not found there.
    Also accepts any single-word match: 'schedule', 'recording', or 'cuts' in the name.
    Prints each folder it checks and any .xlsx files it finds there.
    """
    def _match(name: str) -> bool:
        n = name.lower()
        return n.endswith(".xlsx") and any(
            kw in n for kw in ("recording schedule", "schedule", "cuts")
        )

    folder = course_folder
    for _ in range(5):
        if not folder or folder == os.path.dirname(folder):
            break
        try:
            names = os.listdir(folder)
            xlsx  = [n for n in names if n.endswith(".xlsx")]
            if xlsx:
                print(f"  .xlsx in {os.path.basename(folder)}/: {', '.join(xlsx)}")
            for name in names:
                if _match(name):
                    print(f"  Excel: {name}")
                    return os.path.join(folder, name)
        except OSError:
            pass
        folder = os.path.dirname(folder)
    return None


def to_seconds(value) -> float | None:
    """
    Convert a cut-point cell value to seconds (relative to clip start).
    Returns None for 'end of video' markers.
    """
    if value is None:
        return None
    if isinstance(value, datetime.time):
        # Excel stores "MM:SS" timecodes as datetime.time(MM, SS, 0)
        return value.hour * 60 + value.minute + value.second
    s = str(value).strip().lower()
    if any(x in s for x in ("end", "last", "eof")):
        return None
    parts = s.split(":")
    try:
        if len(parts) == 2:
            return int(parts[0]) * 60 + int(parts[1])
        if len(parts) == 3:
            return int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
    except ValueError:
        pass
    return None


def fmt(secs: float | None) -> str:
    if secs is None:
        return "End of clip"
    m, s = divmod(int(secs), 60)
    return f"{m:02d}:{s:02d}"


def _open_cuts_sheet(excel_path: str):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    for name in wb.sheetnames:
        if name.lower() == "cuts":
            return wb[name]
    raise ValueError("No 'Cuts' sheet found in the Excel file.")


def get_cuts(excel_path: str, chapter_name: str) -> list[tuple[float, float | None]]:
    """Return [(start_secs, end_secs)] for the given chapter. end_secs=None → end of clip."""
    sheet  = _open_cuts_sheet(excel_path)
    target = normalize(chapter_name)
    cuts   = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        if normalize(str(row[0])) == target:
            start = to_seconds(row[1])
            end   = to_seconds(row[2])
            if start is not None:
                cuts.append((start, end))
    return cuts


def get_all_cuts(excel_path: str) -> list[tuple[str, list[tuple[float, float | None]]]]:
    """
    Read the entire Cuts sheet and return one entry per unique chapter name,
    preserving the order they appear in the sheet.
    Returns [(chapter_name, [(start_secs, end_secs), ...]), ...]
    """
    sheet   = _open_cuts_sheet(excel_path)
    ordered: list[str] = []          # chapter names in appearance order
    by_name: dict[str, list] = {}    # chapter_name → cut list

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        name  = str(row[0]).strip()
        start = to_seconds(row[1])
        end   = to_seconds(row[2])
        if start is None:
            continue
        if name not in by_name:
            ordered.append(name)
            by_name[name] = []
        by_name[name].append((start, end))

    return [(n, by_name[n]) for n in ordered]


# ─── JSX generation ───────────────────────────────────────────────────────────
# Two-call architecture:
#   Call 1: generate_split_jsx()  — trim clip.outPoint + seq.overwriteClip (no color)
#   Call 2: generate_color_jsx()  — find clips at cut points, apply label 8 (Mango)
#
# Coloring must be a separate call because nc.label = 8 silently fails on clips
# placed by overwriteClip within the same evalScript call. A 1 s sleep between
# calls lets Premiere's DOM settle so the freshly placed clips become colorable.
#
# QE razor was confirmed to be a no-op on this sequence (fires without error but
# does not create clip boundaries). Pure ExtendScript DOM ops are used instead.

def _cuts_js(cuts: list[tuple[float, float | None]]) -> str:
    rows = []
    for s, e in cuts:
        e_js = "null" if e is None else str(e)
        rows.append(f"  {{start:{s}, end:{e_js}}}")
    return "[\n" + ",\n".join(rows) + "\n]"


def _seq_finder_js(chapter_id: str, part_num: str) -> str:
    """Returns JS preamble that sets seq / seqStart / seqEnd, or returns an error string."""
    return f"""
    function norm(s)  {{ return s.toLowerCase().replace(/[^a-z0-9]/g, ""); }}
    function mt(secs) {{ var t = new Time(); t.seconds = secs; return t; }}
    function matchesPart(cn, pn) {{
        var n = norm(cn), key = "part" + pn, idx = n.indexOf(key);
        if (idx < 0) return false;
        var nx = n.charAt(idx + key.length);
        return nx === "" || isNaN(parseInt(nx, 10));
    }}

    var CHAPTER_ID = "{chapter_id}";
    var PART_NUM   = "{part_num}";

    var seq = null, i;
    for (i = 0; i < app.project.sequences.numSequences; i++)
        if (norm(app.project.sequences[i].name).indexOf(norm(CHAPTER_ID)) === 0)
            {{ seq = app.project.sequences[i]; break; }}
    if (!seq) return "error: seq not found";

    var seqStart = null, seqEnd = null, _t, _c;
    for (_t = 0; _t < seq.videoTracks.numTracks; _t++)
        for (_c = 0; _c < seq.videoTracks[_t].clips.numItems; _c++)
            if (matchesPart(seq.videoTracks[_t].clips[_c].name, PART_NUM)) {{
                var _s = seq.videoTracks[_t].clips[_c].start.seconds;
                var _e = seq.videoTracks[_t].clips[_c].end.seconds;
                if (seqStart === null || _s < seqStart) seqStart = _s;
                if (seqEnd   === null || _e > seqEnd)   seqEnd   = _e;
            }}
    if (seqStart === null) return "error: clip not found";

    try {{ app.project.activeSequence = seq; }} catch(e) {{}}

    // ── Remove existing AutoCut markers for THIS part only ───────────────────
    var _suffix = " Part{part_num}";
    var _mk = seq.markers.getFirstMarker();
    while (_mk !== undefined) {{
        var _next = seq.markers.getNextMarker(_mk);
        var _n = _mk.name || "";
        if ((_n.indexOf("CUT START") === 0 || _n.indexOf("CUT END") === 0 ||
             _n.indexOf("DELETE START") === 0 || _n.indexOf("DELETE END") === 0) &&
            _n.indexOf(_suffix) >= 0)
            seq.markers.deleteMarker(_mk);
        _mk = _next;
    }}
"""


def generate_split_jsx(chapter_id: str, part_num: str, cuts: list[tuple[float, float | None]],
                       color_offset: int = 0) -> str:
    """
    Place a sequence marker at each cut point.
    color_offset: global counter so colours cycle across chapters, not just within one.
    """
    cuts_js = _cuts_js(cuts)
    preamble = _seq_finder_js(chapter_id, part_num)
    return f"""(function () {{
{preamble}
    // setColorByIndex(colorIndex, 0) — correct API (property assignment doesn't work)
    // 0=Green 1=Red 2=Purple 3=Orange 4=Yellow 5=White 6=Blue 7=Cyan
    var COLORS = [0, 1, 6];   // Green / Red / Blue — cycling per pair globally
    var COLOR_OFFSET = {color_offset};
    var cuts  = {cuts_js};
    var added = 0, ki, cut, s0, s1, mk, col;

    for (ki = 0; ki < cuts.length; ki++) {{
        cut = cuts[ki];
        s0  = seqStart + cut.start;
        s1  = (cut.end !== null) ? seqStart + cut.end : seqEnd;
        col = COLORS[(ki + COLOR_OFFSET) % COLORS.length];

        try {{
            mk = seq.markers.createMarker(s0);
            mk.name = "CUT START " + (ki+1) + " Part{part_num}";
            mk.setColorByIndex(col, 0);
            added++;
        }} catch(e) {{ return "error@"+s0.toFixed(1)+": "+e.message; }}

        try {{
            mk = seq.markers.createMarker(s1);
            mk.name = "CUT END " + (ki+1) + " Part{part_num}";
            mk.setColorByIndex(col, 0);
            added++;
        }} catch(e) {{ return "error@"+s1.toFixed(1)+": "+e.message; }}
    }}
    return "v{VERSION}: "+added+" marker(s) col="+col+" @"+seqStart.toFixed(1);
}})();"""


def generate_color_jsx(chapter_id: str, part_num: str, cuts: list[tuple[float, float | None]]) -> str:
    """
    Call 2 of 2 — color Mango (label 8) all clips that start at each cut point.
    Run after a 1 s sleep so Premiere's DOM has settled.
    Scans all video + audio tracks for clips within 0.3 s of the cut start time.
    Also colors clips at the cut END if it doesn't coincide with the clip end.
    """
    cuts_js = _cuts_js(cuts)
    preamble = _seq_finder_js(chapter_id, part_num)
    return f"""(function () {{
{preamble}
    var cuts         = {cuts_js};
    var LABEL_ORANGE = 8;
    var colored      = 0, ki, cut, s0, s1, j, c;

    function colorAt(secs) {{
        var cl;
        for (j = 0; j < seq.videoTracks.numTracks; j++)
            for (c = 0; c < seq.videoTracks[j].clips.numItems; c++) {{
                cl = seq.videoTracks[j].clips[c];
                if (matchesPart(cl.name, PART_NUM) && Math.abs(cl.start.seconds - secs) < 0.3)
                    {{ cl.label = LABEL_ORANGE; colored++; }}
            }}
        for (j = 0; j < seq.audioTracks.numTracks; j++)
            for (c = 0; c < seq.audioTracks[j].clips.numItems; c++) {{
                cl = seq.audioTracks[j].clips[c];
                if (matchesPart(cl.name, PART_NUM) && Math.abs(cl.start.seconds - secs) < 0.3)
                    {{ cl.label = LABEL_ORANGE; colored++; }}
            }}
    }}

    for (ki = 0; ki < cuts.length; ki++) {{
        cut = cuts[ki];
        s0  = seqStart + cut.start;
        s1  = (cut.end !== null) ? seqStart + cut.end : seqEnd;
        colorAt(s0);
        if (Math.abs(s1 - seqEnd) > 0.5) colorAt(s1);
    }}
    return "v{VERSION} color: " + colored + " clips Mango";
}})();"""


# ─── Premiere execution ───────────────────────────────────────────────────────

def get_course_folder_from_premiere() -> str | None:
    """
    Ask Premiere for the media path of any clip in any sequence, then
    walk up the path tree to find the '03_Course' folder and return it.
    Returns None if the server is unreachable or no suitable path is found.
    """
    jsx = """(function() {
        function firstMediaPath(seq) {
            if (!seq) return "";
            for (var t = 0; t < seq.videoTracks.numTracks; t++)
                for (var c = 0; c < seq.videoTracks[t].clips.numItems; c++)
                    try {
                        var p = seq.videoTracks[t].clips[c].projectItem.getMediaPath();
                        // Only accept paths that contain 03_Course (skip assets/backgrounds)
                        if (p && p.indexOf("03_Course") >= 0) return p;
                    } catch(e) {}
            return "";
        }
        // Prefer active sequence; fall back to all sequences in project
        var path = firstMediaPath(app.project.activeSequence);
        if (!path)
            for (var i = 0; i < app.project.sequences.numSequences && !path; i++)
                path = firstMediaPath(app.project.sequences[i]);
        return path;
    })();"""

    try:
        payload = json.dumps({"jsx": jsx}).encode("utf-8")
        req = urllib.request.Request(
            f"{AUTOCUT_SERVER}/run",
            data=payload,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=10) as r:
            body = json.loads(r.read())
            media_path = body.get("result", "")
    except Exception:
        return None

    if not media_path or media_path.startswith("error"):
        return None

    # Walk up the path to find a component named '03_Course'
    parts = media_path.replace("\\", "/").split("/")
    for i, part in enumerate(parts):
        if part == "03_Course":
            return "/".join(parts[: i + 1])

    return None


def ping_server() -> bool:
    try:
        with urllib.request.urlopen(f"{AUTOCUT_SERVER}/ping", timeout=3) as r:
            return r.read() == b"AutoCut OK"
    except Exception:
        return False


def run_in_premiere(jsx_code: str) -> bool:
    if not ping_server():
        print(
            "  ERROR: AutoCut server not reachable on port 7788.\n"
            "  → In Premiere Pro:  Window  ▸  Extensions  ▸  AutoCut\n"
            "    (may need to restart Premiere once after first install)\n"
            "  Then try again."
        )
        return False

    payload = json.dumps({"jsx": jsx_code}).encode("utf-8")
    req = urllib.request.Request(
        f"{AUTOCUT_SERVER}/run",
        data=payload,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=30) as r:
            body = json.loads(r.read())
            result = body.get("result", "")
            if not body.get("success"):
                print(f"  Premiere ERROR: {result}")
                return False
            print(f"  Premiere: {result}")
            return True
    except urllib.error.URLError as e:
        print(f"  HTTP error: {e}")
        return False


# ─── Excel write-back ────────────────────────────────────────────────────────

def _mark_excel_done(excel_path: str, done_chapters: set[str]) -> None:
    """
    Write a ✓ and today's date in column D for every row whose chapter
    (column A) is in done_chapters. Creates a 'Done' header in D1 if missing.
    Skips rows that already have a value in column D (won't overwrite).
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font

    wb    = openpyxl.load_workbook(excel_path)
    sheet = None
    for name in wb.sheetnames:
        if name.lower() == "cuts":
            sheet = wb[name]
            break
    if sheet is None:
        print("  (Excel write-back skipped: no Cuts sheet)")
        return

    # Ensure column D header
    if sheet["D1"].value is None:
        sheet["D1"].value = "Done"
        sheet["D1"].font  = Font(bold=True)

    today      = datetime.date.today().strftime("%Y-%m-%d")
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    updated    = 0

    for row in sheet.iter_rows(min_row=2):
        if not row[0].value:
            continue
        if normalize(str(row[0].value)) in done_chapters:
            d_cell = row[3]           # column D (0-indexed: col A=0, D=3)
            if d_cell.value:          # already marked — don't overwrite
                continue
            d_cell.value = f"✓ {today}"
            d_cell.font  = Font(color="276221")
            d_cell.fill  = green_fill
            updated += 1

    wb.save(excel_path)
    print(f"  Excel updated: {updated} row(s) marked ✓ in column D")


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    single = sys.argv[1] if len(sys.argv) > 1 else None

    # Explicit override as 2nd arg → use it; otherwise auto-detect from Premiere
    if len(sys.argv) > 2:
        course_folder = sys.argv[2]
    else:
        detected = get_course_folder_from_premiere()
        if detected:
            course_folder = detected
            print(f"AutoCut v{VERSION}  (course folder auto-detected from Premiere)")
        else:
            course_folder = DEFAULT_COURSE_FOLDER
            print(f"AutoCut v{VERSION}  (using default course folder)")

    print(f"Course  : {course_folder}\n")

    # ── Find Excel ────────────────────────────────────────────────────────────
    excel_path = find_excel(course_folder)
    if not excel_path:
        print(f"ERROR: No 'Recording Schedule*.xlsx' found near:\n  {course_folder}")
        sys.exit(1)
    print(f"Excel   : {os.path.basename(excel_path)}\n")

    # ── Build work list ───────────────────────────────────────────────────────
    if single:
        cuts = get_cuts(excel_path, single)
        work = [(single, cuts)]
    else:
        work = get_all_cuts(excel_path)

    if not work:
        print("ERROR: No chapters found in the Cuts sheet.")
        sys.exit(1)

    print(f"Chapters to process: {len(work)}\n")

    # ── Ping Premiere once before the loop ────────────────────────────────────
    if not ping_server():
        print(
            "ERROR: AutoCut server not reachable on port 7788.\n"
            "→ In Premiere Pro:  Window ▸ Extensions ▸ AutoCut"
        )
        sys.exit(1)

    # ── Process each chapter ──────────────────────────────────────────────────
    ok = skip = fail = 0
    color_offset  = 0
    done_chapters: set[str] = set()

    for chapter_name, cuts in work:
        chapter_id, part_num = parse_chapter_name(chapter_name)

        if not chapter_id:
            print(f"  SKIP  {chapter_name!r}  (cannot parse Chapter/Part)")
            skip += 1
            continue
        if not cuts:
            print(f"  SKIP  {chapter_name!r}  (no cuts in sheet)")
            skip += 1
            continue

        cut_desc = "  ".join(f"{fmt(s)}→{fmt(e)}" for s, e in cuts)
        print(f"  [{chapter_name}]  {cut_desc}", end="  ", flush=True)

        jsx = generate_split_jsx(chapter_id, part_num, cuts, color_offset)
        payload = json.dumps({"jsx": jsx}).encode("utf-8")
        req = urllib.request.Request(
            f"{AUTOCUT_SERVER}/run",
            data=payload,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        try:
            with urllib.request.urlopen(req, timeout=30) as r:
                body   = json.loads(r.read())
                result = body.get("result", "")
                if not body.get("success"):
                    print(f"→ ERROR: {result}")
                    fail += 1
                elif result.startswith("error:"):
                    print(f"→ {result}")
                    skip += 1
                else:
                    print(f"→ {result}")
                    ok += 1
                    color_offset += len(cuts)
                    done_chapters.add(normalize(chapter_name))
        except Exception as e:
            print(f"→ HTTP error: {e}")
            fail += 1

    print(f"\n{'─'*60}")
    print(f"Done.  ✓ {ok} marked   ⚠ {skip} not in project   ✗ {fail} failed")

    # ── Write ✓ back into the Excel for every successfully processed row ──────
    if done_chapters:
        _mark_excel_done(excel_path, done_chapters)


if __name__ == "__main__":
    main()


# ─── Debug helper ─────────────────────────────────────────────────────────────

def debug_list_clips(chapter_id: str) -> None:
    """Send JSX that lists all clip names in the matching sequence."""
    jsx = f"""
(function(){{
    function norm(s){{ return s.toLowerCase().replace(/[^a-z0-9]/g,""); }}
    var CHAPTER_ID = "{chapter_id}";
    var seq = null;
    for(var i=0;i<app.project.sequences.numSequences;i++){{
        if(norm(app.project.sequences[i].name).indexOf(CHAPTER_ID)===0){{
            seq=app.project.sequences[i]; break;
        }}
    }}
    if(!seq) return "NO SEQ for: "+CHAPTER_ID;
    var out = "Sequence: "+seq.name+"\\n\\nVideo track clips:\\n";
    for(var t=0;t<seq.videoTracks.numTracks;t++){{
        var tr=seq.videoTracks[t];
        if(tr.clips.numItems===0) continue;
        out += "  V"+(t+1)+": ";
        var names=[];
        for(var c=0;c<tr.clips.numItems;c++) names.push(tr.clips[c].name);
        out += names.join(", ")+"\\n";
    }}
    return out;
}})();
"""
    import json, urllib.request
    payload = json.dumps({"jsx": jsx}).encode()
    req = urllib.request.Request("http://127.0.0.1:7788/run", data=payload,
                                  headers={"Content-Type":"application/json"}, method="POST")
    with urllib.request.urlopen(req, timeout=10) as r:
        body = json.loads(r.read())
        print(body.get("result","(no result)"))

if __name__ == "__main__" and len(sys.argv) == 3 and sys.argv[2] == "--debug":
    debug_list_clips(sys.argv[1])
